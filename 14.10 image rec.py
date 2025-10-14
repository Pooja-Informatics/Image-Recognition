import fitz  # PyMuPDF
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.widgets import RectangleSelector
import easyocr
import re
import cv2
import os
from tkinter import Tk, filedialog
from openpyxl import Workbook

# -------------------- Config --------------------
DEBUG = False  # set True to see extra debug prints
# -------------------- Select PDF --------------------
def select_pdf_file():
    root = Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(
        title="Select a PDF file",
        filetypes=[("PDF files", "*.pdf")]
    )
    return file_path

# -------------------- Pixmap to OpenCV --------------------
def pix_to_cv2(pix):
    img_bytes = pix.tobytes("ppm")
    img_np = cv2.imdecode(np.frombuffer(img_bytes, np.uint8), cv2.IMREAD_COLOR)
    return img_np

# -------------------- Preprocess Image for OCR --------------------
def preprocess_image(img):
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

    # upscale to help OCR
    h, w = gray.shape
    scale = 2 if max(h, w) < 2000 else 1  # upscale smaller images
    if scale != 1:
        gray = cv2.resize(gray, None, fx=scale, fy=scale, interpolation=cv2.INTER_LINEAR)

    # gentle blur to reduce noise then adaptive threshold
    blurred = cv2.GaussianBlur(gray, (3, 3), 0)
    thresh = cv2.adaptiveThreshold(blurred, 255,
                                   cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                                   cv2.THRESH_BINARY, 11, 2)

    # morphological close to join broken characters
    kernel = np.ones((2, 2), np.uint8)
    thresh = cv2.morphologyEx(thresh, cv2.MORPH_CLOSE, kernel)
    return thresh

# -------------------- Multi-Crop Tool (Zoom + Selection) --------------------
def get_multiple_crop_coords(image):
    coords_list = []
    xlim, ylim = [0, image.shape[1]], [0, image.shape[0]]

    def onselect(eclick, erelease):
        x1, y1 = int(eclick.xdata), int(eclick.ydata)
        x2, y2 = int(erelease.xdata), int(erelease.ydata)
        rect = {
            'x': max(0, min(x1, x2)),
            'y': max(0, min(y1, y2)),
            'w': abs(x2 - x1),
            'h': abs(y2 - y1)
        }
        coords_list.append(rect)
        print(f"📌 Added rectangle: {rect}")

    def onscroll(event):
        nonlocal xlim, ylim
        if event.inaxes != ax:
            return
        cur_x, cur_y = event.xdata, event.ydata
        if cur_x is None or cur_y is None:
            return
        scale = 0.8 if event.button == 'up' else 1.2 if event.button == 'down' else 1
        w = (xlim[1] - xlim[0]) * scale
        h = (ylim[1] - ylim[0]) * scale
        xlim = [cur_x - w/2, cur_x + w/2]
        ylim = [cur_y - h/2, cur_y + h/2]
        ax.set_xlim(xlim)
        ax.set_ylim(ylim[::-1])
        fig.canvas.draw_idle()

    fig, ax = plt.subplots(figsize=(12, 10))
    ax.imshow(cv2.cvtColor(image, cv2.COLOR_BGR2RGB))
    ax.set_title("Draw boxes | Press 'q' to finish | Scroll to zoom")

    rect_selector = RectangleSelector(
        ax, onselect, useblit=True,
        button=[1], minspanx=5, minspany=5,
        interactive=True
    )

    def on_key(event):
        if event.key.lower() == 'q':
            rect_selector.set_active(False)
            plt.close()

    fig.canvas.mpl_connect("key_press_event", on_key)
    fig.canvas.mpl_connect("scroll_event", onscroll)
    plt.show()
    return coords_list

# -------------------- Robust Number Extraction --------------------last code
import re
 
DEBUG = False
ENABLE_CORRECTIONS = True
 
def correct_ocr_errors(token):
    token = token.replace("|", "1").replace("I", "1").replace("l", "1").replace("!", "1").replace("/", "1").replace("O", "0")

    # Fix short dashed suffix
    if "-" in token:
        base, suffix = token.split("-")
        if len(suffix) == 2:
            token = base + "-001"
        elif len(suffix) == 3 and not suffix.startswith("1"):
            token = base + "-1" + suffix

    # Fix 5-digit base + letters: 20302AA → 203012AA
    m = re.match(r"^(\d{5})([A-Z]{2,3})$", token)
    if m:
        base, suffix = m.groups()
        corrected = base[:4] + "1" + base[4:] + suffix
        return corrected

    # ✅ Truncate bad alphanum suffixes (like ESE → E)
    m2 = re.match(r"^(\d{6})([A-Z]{2,})$", token)
    if m2:
        base, suffix = m2.groups()
        token = base + suffix[:1]

    return token

def extract_text_and_numbers(ocr_texts):
    full = " ".join([t.strip() for t in ocr_texts if t and t.strip()])
    if not full:
        return []

    # Normalize
    full = re.sub(r"[–—−‐]", "-", full)
    full = re.sub(r"\s*-\s*", "-", full)
    full = re.sub(r"\s+", " ", full).strip()
    full_upper = full.upper()

    digit_string = re.sub(r"[^A-Z0-9-]", "", full_upper.replace(" ", ""))

    candidates = []

    # R-prefixed
    for m in re.finditer(r"\bR\W{0,2}\d{6,9}(?:\W{0,2}-\W{0,2}\d{1,4})?", full_upper):
        candidates.append(re.sub(r"[^A-Z0-9-]", "", m.group(0)))

    # Dashed long
    candidates += re.findall(r"\d{6,9}-\d{1,4}", digit_string)

    # Short dashed
    candidates += re.findall(r"\d{2,4}-\d{3,6}", digit_string)

    # Alphanumeric
    candidates += re.findall(r"\d{5,9}[A-Z0-9]{1,3}", digit_string)

    # Deduplicate
    candidates = list(dict.fromkeys(candidates))

    cleaned = []
    seen_cleaned = set()
    for token in candidates:
        token = re.sub(r"[^A-Z0-9-]", "", token).strip("-")
        if not token or token.count("-") > 1:
            continue

        # Apply OCR correction
        if ENABLE_CORRECTIONS:
            token = correct_ocr_errors(token)

        # Fix known OCR issues
        if token.startswith("2500"):
            token = "5" + token[2:]

        if "-" in token:
            base, *rest = token.split("-")
            if len(base) >= 9 and base[0] in {"2", "3"} and base[1:4] == "500":
                base = base[1:]
                token = "-".join([base] + rest)

        if token.startswith(("21", "22")):
            if "-" in token:
                parts = token.split("-")
                if len(parts[0]) > 8:
                    token = parts[0][1:] + "-" + parts[1]
            else:
                digits_only = re.sub(r"\D", "", token)
                suffix = re.sub(r"\d", "", token)
                if len(digits_only) > 6:
                    token = digits_only[1:] + suffix

        if token.startswith("R") and not re.match(r"^R\d{6,9}(?:-\d{1,4})?$", token):
            continue

        if token not in seen_cleaned:
            cleaned.append(token)
            seen_cleaned.add(token)

    # Merge R-prefixed and plain part numbers
    r_by_base = {}
    p_by_base = {}
    others = []

    def split_core(t):
        m = re.match(r"^(R)?(\d{6,9})(?:-(\d{1,4}))?$", t)
        if m:
            return m.group(1) or "", m.group(2), m.group(3) or ""
        return None, None, None

    for t in cleaned:
        pref, base, suf = split_core(t)
        if base is None:
            others.append(t)
            continue
        if pref == "R":
            if base not in r_by_base or len(suf) > len(r_by_base[base]):
                r_by_base[base] = suf
        else:
            if base not in p_by_base or len(suf) > len(p_by_base[base]):
                p_by_base[base] = suf

    merged = []
    bases = set(r_by_base) | set(p_by_base)
    for base in bases:
        suf_r = r_by_base.get(base, "")
        suf_p = p_by_base.get(base, "")
        best_suf = suf_r if len(suf_r) >= len(suf_p) else suf_p
        if base in r_by_base:
            out = "R" + base + (f"-{best_suf}" if best_suf else "")
        else:
            out = base + (f"-{best_suf}" if best_suf else "")
        merged.append(out)

    def score(t):
        digits_only = re.sub(r"\D", "", t)
        return (
            1 if t.startswith("R") else 0,
            1 if "-" in t else 0,
            len(digits_only),
            len(t),
        )

    kept = sorted(merged, key=score, reverse=True)
    for tok in sorted(others, key=score, reverse=True):
        if any(tok == m or tok in m or m in tok for m in kept):
            continue
        kept.append(tok)

    final = []
    for tok in sorted(kept, key=score, reverse=True):
        if any(tok != k and tok in k for k in final):
            continue
        if tok not in final:
            final.append(tok)

    # Extract descriptive text
    name_text = re.sub(r"[A-Z]?\d+(?:-\d+)?[A-Z0-9]{0,3}", " ", full_upper)
    name_text = re.sub(r"\s+", " ", name_text).strip()
    name_text = name_text if re.search(r"[A-Z]{2,}", name_text) else ""

    return [(num, name_text) for num in final] if final else ([("", name_text)] if name_text else [])

# -------------------- Save to Excel --------------------
def save_to_excel(pairs, pdf_path):
    os.makedirs("files", exist_ok=True)
    base_name = os.path.splitext(os.path.basename(pdf_path))[0]
    filename = os.path.join("files", base_name + ".xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted Data"

    ws["A1"] = "Part Numbers"
    ws["B1"] = "Part Names"

    row = 2
    for num, text in pairs:
        if num:
            cell = ws.cell(row=row, column=1, value=str(num))
            # force text format to avoid Excel auto-date conversion
            cell.number_format = '@'
        if text and not re.search(r'\d', text):
            ws.cell(row=row, column=2, value=text)
        row += 1

    wb.save(filename)
    print(f"📊 Data saved to {filename}")

# -------------------- Main Logic --------------------
def main():
    pdf_path = select_pdf_file()
    if not pdf_path:
        print("❌ No file selected.")
        return

    doc = fitz.open(pdf_path)
    # Use cpu by default (safer). Change gpu=True if you have a CUDA GPU and configured easyocr.
    reader = easyocr.Reader(['en'], gpu=False)

    all_pairs = []

    for page_num in range(len(doc)):
        page = doc.load_page(page_num)
        matrix = fitz.Matrix(4, 4)
        pix = page.get_pixmap(matrix=matrix, alpha=False)
        img_cv = pix_to_cv2(pix)

        print(f"\n📄 Page {page_num + 1} — Draw rectangles (press 'q' when done).")
        crop_coords_list = get_multiple_crop_coords(img_cv)
        if not crop_coords_list:
            print("⚠️ No regions selected. Skipping this page.")
            continue

        for i, crop_coords in enumerate(crop_coords_list, start=1):
            x, y, w, h = crop_coords['x'], crop_coords['y'], crop_coords['w'], crop_coords['h']
            if w <= 0 or h <= 0:
                print(f"⚠️ Skipping Box {i} (invalid size).")
                continue

            pad = 5
            cropped = img_cv[max(0, y-pad):y+h+pad, max(0, x-pad):x+w+pad]
            if cropped.size == 0:
                print(f"⚠️ Skipping Box {i} (empty crop).")
                continue

            preprocessed = preprocess_image(cropped)

            # OCR: we get a list of text snippets
            ocr_result = reader.readtext(preprocessed, detail=0)
            pairs = extract_text_and_numbers(ocr_result)

            if not pairs:
                print(f"⚠️ Page {page_num + 1}, Box {i}: no valid part numbers found. OCR output: {ocr_result}")
            else:
                print(f"✅ Page {page_num + 1}, Box {i} extracted:")
                for num, txt in pairs:
                    print(f"   Number → {num}, Text → {txt}")

            all_pairs.extend(pairs)

    if not all_pairs:
        print("⚠️ No pairs were extracted from the document.")
    else:
        save_to_excel(all_pairs, pdf_path)

# -------------------- Run --------------------
if __name__ == "__main__":
    main()
